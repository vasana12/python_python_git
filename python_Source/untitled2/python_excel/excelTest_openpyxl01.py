#http://openpyxl.readthedocs.io/en/stable/ 사이트 참조
#설치 : pip install openpyxl (file -> settings-> + 버튼 누르고 openpyxl 인스톨)

from openpyxl import Workbook, load_workbook

wb = Workbook() #엑셀문서에서 새 문서를 만드는 것과 동일한 역할 자동으로 워크시트가 하나 만들어진다.
ws = wb.active # 활성화 되어 잇는 시트 선택
#ws = wb.create_sheet(0) # 0 이면 첫번째 위치에 삽입 ,생략시 마지막 위치에 삽입
ws.title = "sample"     #워크시트 이름 변경
ws['B2'] = 42
ws.append([1,2,3])      #워크시트에서 가장 마지막에 추가된 데이터 행 그다음 행 부터 추가한다.
ws.append([1,2,3])
ws.append([1,2,3])
ws.append([1,2,3])
ws.append([4,4,4])
wb.save("openpyxl1.xlsx")   #파일생성
wb.close()                  #파일 닫아주기

wb=load_workbook(filename= 'openpyxl1.xlsx')    #데이터 불러오는 과정
ws = wb.active
ws['A1'] =42
ws.cell(row=1, column=3).value=333 #1행 3열 위치에 대입하라
ws.append(['aaa','bbb','ccc']) # 새로운 행에 추가됨

print(ws['A1'].value)
print(ws['A2'].value) #저장된 값이 없으면 None으로 출력

ws2 = wb['sample'] # sample 시트 선택
print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)
wb.save("openpyxl1.xlsx")