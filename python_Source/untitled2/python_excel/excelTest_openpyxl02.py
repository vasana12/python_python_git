from openpyxl import Workbook

wb = Workbook() #엑셀문서에서 새 문서를 만드는 것과 동일한 역할 자동으로 워크시트가 하나 만들어진다.
ws = wb.active  # 활성화 되어 잇는 시트 선택
ws = wb.create_sheet("diary",0) #0 이면 첫번째 위치에 삽입, 생략시 마지막 위치에 삽입

data = [('홍길동',80,70,90),('이기자',90,60,80),('강기자',80,80,70)]

r =1 #첫번째 행이 1부터 시작함
c =1 #첫번째 열이 1부터 시작함

for irum, kor, eng, math in data:
    ws.cell(row=r, column=c).value = irum
    ws.cell(row=r, column=c+1).value = kor
    ws.cell(row=r, column=c+2).value = eng
    ws.cell(row=r, column=c+3).value = math
    r +=1
"""
for irum, kor, eng, math in data:
    ws['A' + str(r)].value = irum
    ws['B' + str(r)].value = kor
    ws['C' + str(r)].value = eng
    ws['D' + str(r)].value = math
    r+=1
    
    
columnChar = 'A'
for irum, kor, eng, math in data:
    ws[columnChar + str(r)].value = irum
    ws[columnChar + str(r)].value = irum.encode(encoding='utf-8',errors='ignore')
    ws[chr(ord(columnChar)+1)+str(r)].value = kor    #ord 는 해당 문자에 대한 코드값 A 에 1 더하면 B가됨
    ws[chr(ord(columnChar)+2)+str(r)].value = eng
    ws[chr(ord(columnChar)+3)+str(r)].value = math
    r+=1
    
    
coulumnChar = 'A'
for val in data:
    for i in range(len()):
        ws[chr(ord(columnChar)+i) + str(r)].value = val[i]
    r+=1
"""

ws.cell(row=r, column=1).value = '합계'
ws.cell(row=r, column=2).value = '=sum(B1:B3)'
ws.cell(row=r, column=3).value = '=sum(C1:C3)'
ws.cell(row=r, column=4).value = '=sum(D1:D3)'

wb.save("openpyxl2.xlsx")
wb.close()